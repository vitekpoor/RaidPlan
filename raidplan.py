#!/usr/bin/env python3
"""
raidplan_rename.py — swap player names in a RaidPlan.io plan across all steps.

Takes a plan URL and a set of renames (Old=New). Clones the plan via the
RaidPlan API, replaces the names in every marker and text label on every
slide, syncs class icons/colors from the guild roster Google Sheet (cell
fill color = WoW class color), saves the clone and prints its URL.
The original plan is never touched.

Usage:
  python raidplan_rename.py <plan-url-or-code> "Old=New" ["Old2=New2" ...]
      [--sheet URL] [--gid N] [--name "Plan name"] [--dry-run] [--check]

Examples:
  python raidplan_rename.py https://raidplan.io/plan/kdwu3c3tfgv5tkdu/edit "Irdwy=Meslock"
  python raidplan_rename.py kdwu3c3tfgv5tkdu "Irdwy=Meslock" "Hase=Mikyywar" --dry-run

Rename separators accepted: "Old=New", "Old->New", "Old - New".
Requires: openpyxl (pip install openpyxl)
"""

import argparse
import io
import json
import re
import sys
import unicodedata
import urllib.request

DEFAULT_SHEET = "https://docs.google.com/spreadsheets/d/1XwIuLB7o0kHDiViG9tVXF1Sk7ZBgvYqh3icINuNqFbE"
DEFAULT_GID = "384829141"

# Standard WoW class colors as used by the roster sheet cell fills.
CLASS_COLORS = {
    "warrior":     (0xC7, 0x9C, 0x6E),
    "paladin":     (0xF5, 0x8C, 0xBA),
    "hunter":      (0xAB, 0xD4, 0x73),
    "rogue":       (0xFF, 0xF5, 0x69),
    "priest":      (0xFF, 0xFF, 0xFF),
    "deathknight": (0xC4, 0x1F, 0x3B),
    "shaman":      (0x00, 0x70, 0xDE),
    "mage":        (0x69, 0xCC, 0xF0),
    "warlock":     (0x94, 0x82, 0xC9),
    "monk":        (0x00, 0xFF, 0x96),
    "druid":       (0xFF, 0x7D, 0x0A),
    "demonhunter": (0xA3, 0x30, 0xC9),
    "evoker":      (0x33, 0x93, 0x7F),
}
# Hex border colors RaidPlan puts on class markers (same palette).
CLASS_HEX = {cls: "#%02X%02X%02X" % rgb for cls, rgb in CLASS_COLORS.items()}

# Spec implied by (class, roster role) — only where the role pins it down.
# Ambiguous combos (Melee warrior, Ranged mage, ...) keep the base class icon.
SPEC_BY_ROLE = {
    "deathknight": {"tank": "blood"},
    "demonhunter": {"tank": "vengeance", "melee": "havoc"},
    "druid":       {"tank": "guardian", "heal": "restoration",
                    "ranged": "balance", "melee": "feral"},
    "evoker":      {"heal": "preservation", "ranged": "devastation"},
    "hunter":      {"melee": "survival"},
    "monk":        {"tank": "brewmaster", "heal": "mistweaver", "melee": "windwalker"},
    "paladin":     {"tank": "protection", "heal": "holy", "melee": "retribution"},
    "priest":      {"heal": "holy", "ranged": "shadow"},
    "shaman":      {"heal": "restoration", "ranged": "elemental", "melee": "enhancement"},
    "warrior":     {"tank": "protection"},
}

ASSUMED_TANK_ROLES = set()


def role_key(role):
    """Map a roster role label to tank/heal/ranged/melee. Unrecognized
    non-empty labels (guild slang like 'Firer', 'Kúň') are assumed tanks."""
    r = norm(role)
    if not r:
        return None
    if "heal" in r:
        return "heal"
    if "range" in r or r in ("rdps", "caster"):
        return "ranged"
    if "melee" in r or r == "mdps":
        return "melee"
    if r != "tank":
        ASSUMED_TANK_ROLES.add(role)
    return "tank"


def spec_asset(cls, role):
    """Asset path with the role-implied spec, or None when ambiguous."""
    rk = role_key(role)
    spec = SPEC_BY_ROLE.get(cls, {}).get(rk) if rk else None
    return f"game/wow/class/{cls}_{spec}.png" if spec else None


def http_json(url, payload=None, headers=None):
    data = json.dumps(payload).encode() if payload is not None else None
    req = urllib.request.Request(
        url, data=data, method="POST" if data else "GET",
        headers={"Content-Type": "application/json", "User-Agent": "raidplan-rename/1.0",
                 **(headers or {})},
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read().decode())


def http_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": "raidplan-rename/1.0"})
    with urllib.request.urlopen(req) as r:
        return r.read()


def norm(name):
    """Lowercase and strip diacritics so 'Anál'/'Drästic' match 'Anal'/'Drastic'."""
    s = unicodedata.normalize("NFKD", name or "")
    return "".join(c for c in s if not unicodedata.combining(c)).strip().lower()


def plan_code(url_or_code):
    """Return (code, key). key is set when a private edit link
    raidplan.io/plan/<code>/<key> was pasted, else None."""
    m = re.search(r"raidplan\.io/plan/([a-z0-9]+)(?:/([A-Za-z0-9]{16,}))?", url_or_code)
    if m:
        return m.group(1), m.group(2)
    return url_or_code.strip().strip("/"), None


def nearest_class(rgb_hex):
    """Map an ARGB/RGB hex fill to the closest WoW class color, or None."""
    h = (rgb_hex or "").lstrip("#")
    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return None
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    best, dist = None, None
    for cls, (cr, cg, cb) in CLASS_COLORS.items():
        d = (r - cr) ** 2 + (g - cg) ** 2 + (b - cb) ** 2
        if dist is None or d < dist:
            best, dist = cls, d
    return best if dist is not None and dist < 3600 else None  # ~60 per channel


def load_roster(sheet_url, gid):
    """Return {normalized_name: (display_name, class, role)} from the Google Sheet.

    Downloads the whole workbook as xlsx (for cell fill colors) plus the target
    tab as CSV (xlsx export loses gids), then picks the xlsx tab whose names
    overlap the CSV the most.
    """
    import openpyxl

    doc_id = re.search(r"/d/([\w-]+)", sheet_url).group(1)
    base = f"https://docs.google.com/spreadsheets/d/{doc_id}/export"
    csv_text = http_bytes(f"{base}?format=csv&gid={gid}").decode("utf-8", "replace")
    csv_names = {norm(line.split(",")[1]) for line in csv_text.splitlines()[1:]
                 if line.count(",") and line.split(",")[1].strip()}

    wb = openpyxl.load_workbook(io.BytesIO(http_bytes(f"{base}?format=xlsx")))
    ws = max(wb.worksheets, key=lambda w: sum(
        1 for row in w.iter_rows(min_col=2, max_col=2) if norm(str(row[0].value or "")) in csv_names))

    roster = {}
    for row in ws.iter_rows(min_col=1, max_col=2):
        role_cell, name_cell = row
        name = name_cell.value
        if not isinstance(name, str) or not name.strip() or name.startswith("="):
            continue
        fill = name_cell.fill
        rgb = None
        if fill and fill.patternType and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb
        cls = nearest_class(rgb) if rgb else "priest"  # uncolored/white cell = priest
        roster[norm(name)] = (name.strip(), cls, str(role_cell.value or "").strip())
    return roster, ws.title


def roster_lookup(roster, name):
    """Exact normalized match first, then prefix match (Nesfe ~ Nesferit)."""
    n = norm(name)
    if n in roster:
        return roster[n]
    if len(n) >= 4:
        hits = [v for k, v in roster.items() if k.startswith(n) or n.startswith(k)]
        if len(hits) == 1:
            return hits[0]
    return None


def replace_in_text(text, renames, hits=None):
    """Word-boundary, case/diacritics-insensitive replacement in free text.

    All names are replaced in ONE pass so swaps (A->B while B->A) don't
    chase each other's output."""
    if not text or not renames:
        return text, 0
    # match against the diacritics-stripped shadow so 'Zirael' finds 'Zîrael'
    shadow = "".join(c for c in unicodedata.normalize("NFKD", text)
                     if not unicodedata.combining(c))
    alts = "|".join(re.escape(o) for o in sorted(renames, key=len, reverse=True))
    pattern = re.compile(rf"(?<!\w)({alts})(?!\w)", re.IGNORECASE)
    out, pos, count = [], 0, 0
    for m in pattern.finditer(shadow):
        old = next(o for o in renames if norm(o) == norm(m.group(1)))
        out.append(text[pos:m.start()]); out.append(renames[old]); pos = m.end(); count += 1
        if hits is not None:
            hits[old] = hits.get(old, 0) + 1
    out.append(text[pos:])
    return "".join(out), count


def key_help(code):
    return f"""\
How to get the access key for plan '{code}' (needed for --in-place):

  The key was saved in the browser where the plan was created/cloned.

  1. Browser console (fastest):
     - open any raidplan.io page, press F12 -> Console tab
     - paste this and press Enter (if Chrome refuses, type 'allow pasting' first):
         JSON.parse(localStorage.getItem('rp:planKeys')).state.keys.find(k => k[0] === '{code}')
     - the second string in the result is the key.

  2. Manually: F12 -> Application tab -> Storage -> Local Storage ->
     https://raidplan.io -> row 'rp:planKeys' -> find "{code}"
     in the value; the 32-character string next to it is the key.

  3. Browser history: Ctrl+H, search for '{code}' - a URL like
     raidplan.io/plan/{code}/<key> has the key as its last part.

  Then run:
    python raidplan_rename.py {code} "Old=New" --in-place --key <the-key>

  Shortcut: paste the full edit link as the plan argument and skip both
  --in-place and --key:
    python raidplan_rename.py https://raidplan.io/plan/{code}/<the-key> "Old=New"

  If the plan belongs to a logged-in RaidPlan account instead, pass the
  site's Cookie header via --cookie (F12 -> Network -> any raidplan.io
  request -> Request Headers -> Cookie).

  Treat the key like a password: anyone who has it can edit the plan."""


def main():
    ap = argparse.ArgumentParser(
        description="Rename players in a RaidPlan.io plan.",
        epilog=key_help("<plan-code>"),
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("plan", help="plan URL or code")
    ap.add_argument("renames", nargs="*",
                    help='renames like "Old=New" / "Old->New" / "Old - New", '
                         'or position swaps like "PlayerA<->PlayerB"')
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help="roster Google Sheet URL")
    ap.add_argument("--gid", default=DEFAULT_GID, help="roster tab gid")
    ap.add_argument("--name", default=None, help="name for the new plan (default: '<old> v2')")
    ap.add_argument("--dry-run", action="store_true", help="show changes, don't create a plan")
    ap.add_argument("--check", action="store_true", help="also report roster/class mismatches for all markers")
    ap.add_argument("--file", default=None,
                    help="read renames from a text file, one 'Old=New' (or 'Old -> New' / "
                         "'Old - New') per line; blank lines and lines starting with # are ignored")
    ap.add_argument("--sync-names", action="store_true",
                    help="rename all plan markers to the exact roster spelling from the sheet "
                         "(matching is case/diacritics-insensitive with prefix fallback)")
    ap.add_argument("--normalize", action="store_true",
                    help="make all player-name markers uniform: icon scale 0.28, name font 12, "
                         "most common label background (override with --marker-scale/--marker-font)")
    ap.add_argument("--marker-scale", type=float, default=None,
                    help="with --normalize: icon scale for player markers (default 0.28)")
    ap.add_argument("--marker-font", type=int, default=None,
                    help="with --normalize: name font size for player markers (default 12)")
    ap.add_argument("--in-place", action="store_true",
                    help="save back to the SAME plan instead of cloning (needs --key or --cookie)")
    ap.add_argument("--key", default="",
                    help="plan access key (browser DevTools -> Application -> Local Storage -> "
                         "raidplan.io -> rp:planKeys, or the tail of a /plan/<code>/<key> edit link)")
    ap.add_argument("--cookie", default="",
                    help="raidplan.io Cookie header value, for plans owned by a logged-in account")
    args = ap.parse_args()

    pairs = list(args.renames)
    if args.file:
        with open(args.file, encoding="utf-8-sig") as f:
            pairs += [line.strip() for line in f
                      if line.strip() and not line.lstrip().startswith("#")]

    renames = {}
    for pair in pairs:
        pair = pair.strip().strip("[](){}\"'")
        sw = re.split(r"\s*(?:<->|<>)\s*", pair, maxsplit=1)
        if len(sw) == 2 and sw[0].strip() and sw[1].strip():
            a, b = (s.strip().strip("[](){}\"'") for s in sw)
            renames[a] = b
            renames[b] = a
            continue
        m = re.split(r"\s*(?:=|->|—|(?<=\S)\s+-\s+(?=\S))\s*", pair, maxsplit=1)
        if len(m) != 2 or not m[0].strip() or not m[1].strip():
            sys.exit(f"Cannot parse rename {pair!r} — use Old=New")
        old, new = (s.strip().strip("[](){}\"'") for s in m)
        if not old or not new:
            sys.exit(f"Cannot parse rename {pair!r} — use Old=New")
        renames[old] = new
    if not renames and not args.check and not args.normalize and not args.sync_names:
        sys.exit("No renames given (and no --check/--normalize/--sync-names) — nothing to do.")

    code, url_key = plan_code(args.plan)
    if url_key and not args.key:
        args.key = url_key
        if not args.in_place:
            args.in_place = True
            print("Edit link detected — will save changes back to this plan (in place).")
    print(f"Fetching plan {code} ...")
    meta = http_json(f"https://raidplan.io/api/plan/{code}")
    plan = meta["plan"]
    doc = http_json(plan["doc_url"])
    original_doc = json.loads(json.dumps(doc))  # pristine copy for --in-place backup
    print(f"  '{plan['name']}' — {plan['title']}, {plan['steps']} steps, {len(doc['nodes'])} nodes")

    print("Loading roster sheet ...")
    roster, tab = load_roster(args.sheet, args.gid)
    print(f"  tab '{tab}': {len(roster)} raiders")

    # Warn early about unknown replacement names, and resolve their classes.
    new_class = {}
    for old, new in renames.items():
        hit = roster_lookup(roster, new)
        if hit:
            new_class[old] = (hit[1], hit[2])
            print(f"  {old} -> {new}: {new} is {hit[1]} ({hit[2]}) in the roster")
        else:
            print(f"  WARNING: '{new}' not found in roster — icon/color will be kept from '{old}'")

    if args.sync_names:
        plan_names = {}
        for node in doc["nodes"]:
            if node["type"] == "marker":
                txt = (node["attr"].get("text") or "").strip()
                if txt:
                    plan_names.setdefault(norm(txt), set()).add(txt)
        print("Syncing marker names to roster spelling:")
        n_sync = 0
        for variants in sorted(plan_names.values(), key=lambda v: min(v).lower()):
            if any(norm(o) == norm(min(variants)) for o in renames):
                continue  # explicit Old=New from the command line wins
            hit = roster_lookup(roster, min(variants))
            if hit is None:
                print(f"  ?  '{min(variants)}' has no roster match - left as is")
                continue
            wrong = sorted(v for v in variants if v != hit[0])
            if wrong:
                renames[wrong[0]] = hit[0]
                new_class.setdefault(wrong[0], (hit[1], hit[2]))
                print(f"  ~  {' / '.join(wrong)} -> {hit[0]}  ({hit[1]}, {hit[2]})")
                n_sync += 1
        if not n_sync:
            print("  all marker names already match the sheet")

    marker_hits = text_hits = icon_hits = 0
    hits = {}
    unknown = set()
    mismatch = set()
    class_fixed = set()
    class_wrong = set()
    for node in doc["nodes"]:
        attr = node.get("attr", {})
        if node["type"] == "marker":
            txt = attr.get("text") or ""
            if txt:
                hit_key = next((o for o in renames if norm(o) == norm(txt)), None)
                if hit_key:
                    if attr["text"] != renames[hit_key]:
                        attr["text"] = renames[hit_key]
                        marker_hits += 1
                    hits[hit_key] = hits.get(hit_key, 0) + 1
                    info = new_class.get(hit_key)
                    if info:
                        cls, role = info
                        old_asset = attr.get("asset") or ""
                        m = re.match(r"game/wow/class/([a-z]+)(?:_([a-z]+))?\.png", old_asset)
                        if not m or m.group(1) != cls:
                            attr["asset"] = spec_asset(cls, role) or f"game/wow/class/{cls}.png"
                            if attr.get("border"):
                                attr["border"] = CLASS_HEX[cls]
                            icon_hits += 1
                        elif not m.group(2) and spec_asset(cls, role):
                            attr["asset"] = spec_asset(cls, role)
                            icon_hits += 1
                elif args.check or args.sync_names:
                    hit = roster_lookup(roster, txt)
                    if not hit:
                        unknown.add(txt)
                    else:
                        if hit[0] != txt:
                            mismatch.add((txt, hit[0]))
                        cls, role = hit[1], hit[2]
                        m = re.match(r"game/wow/class/([a-z]+)(?:_([a-z]+))?\.png",
                                     attr.get("asset") or "")
                        if m and m.group(1) != cls:
                            if args.sync_names:
                                attr["asset"] = spec_asset(cls, role) or f"game/wow/class/{cls}.png"
                                if attr.get("border"):
                                    attr["border"] = CLASS_HEX[cls]
                                icon_hits += 1
                                class_fixed.add((txt, m.group(1),
                                                 attr["asset"].rsplit("/", 1)[-1][:-4]))
                            else:
                                class_wrong.add((txt, m.group(1), cls))
                        elif m and not m.group(2) and args.sync_names:
                            new_asset = spec_asset(cls, role)
                            if new_asset:
                                attr["asset"] = new_asset
                                icon_hits += 1
                                class_fixed.add((txt, m.group(1),
                                                 new_asset.rsplit("/", 1)[-1][:-4]))
        elif node["type"] == "itext":
            attr["text"], n = replace_in_text(attr.get("text"), renames, hits)
            text_hits += n

    for key in ("header_notes_raw", "footer_notes_raw"):
        doc[key], n = replace_in_text(doc.get(key) or "", renames, hits)
        text_hits += n
    step_notes = []
    for note in doc.get("step_notes_raw") or []:
        note, n = replace_in_text(note or "", renames, hits)
        text_hits += n
        step_notes.append(note)

    fmt_hits = 0
    if args.normalize:
        from collections import Counter
        named = [n for n in doc["nodes"]
                 if n["type"] == "marker" and (n["attr"].get("text") or "").strip()]
        if named:
            t_scale = args.marker_scale or 0.28
            t_font = args.marker_font or 12
            t_bg = Counter(n["attr"].get("labelBackground") for n in named).most_common(1)[0][0]
            print(f"\nNormalizing {len(named)} player markers to scale {t_scale}, "
                  f"font {t_font}, label background {t_bg or 'none'} ...")
            for n in named:
                changed = (round(n["meta"]["scale"]["x"], 2) != t_scale
                           or round(n["meta"]["scale"]["y"], 2) != t_scale
                           or n["attr"].get("fontSize") != t_font
                           or n["attr"].get("labelBackground") != t_bg)
                if changed:
                    n["meta"]["scale"]["x"] = n["meta"]["scale"]["y"] = t_scale
                    n["attr"]["fontSize"] = t_font
                    n["attr"]["labelBackground"] = t_bg
                    fmt_hits += 1
            print(f"  {fmt_hits} markers reformatted, {len(named) - fmt_hits} already matched.")

    print(f"\nChanged {marker_hits} player markers, {text_hits} mentions in texts/notes, "
          f"{icon_hits} class icons updated" +
          (f", {fmt_hits} markers reformatted." if args.normalize else "."))
    for old in renames:
        if not hits.get(old):
            print(f"  WARNING: '{old}' does not appear anywhere in the plan — check the spelling")
    if args.check and unknown:
        print("Markers whose name is NOT in the roster sheet:")
        for t in sorted(unknown):
            print(f"  - {t}")
    if args.check and mismatch:
        print("Markers spelled differently than the roster sheet (fix with --sync-names):")
        for t, r in sorted(mismatch):
            print(f"  - {t}  (sheet: {r})")
    if class_fixed:
        print("Class/spec icons synced to the roster sheet:")
        for t, old_cls, cls in sorted(class_fixed):
            print(f"  - {t}: {old_cls} -> {cls}")
    if ASSUMED_TANK_ROLES:
        print(f"  (roles {', '.join(repr(r) for r in sorted(ASSUMED_TANK_ROLES))} "
              f"were treated as tank)")
    if args.check and class_wrong:
        print("Markers whose class differs from the roster sheet (fix with --sync-names):")
        for t, old_cls, cls in sorted(class_wrong):
            print(f"  - {t}: plan {old_cls}, sheet {cls}")
    if args.check:
        if not unknown and not mismatch and not class_wrong:
            print("All plan names and classes match the roster sheet.")
        matched = set()
        for node in doc["nodes"]:
            if node["type"] == "marker":
                txt = (node["attr"].get("text") or "").strip()
                hit = roster_lookup(roster, txt) if txt else None
                if hit:
                    matched.add(hit[0])
        missing = [v for _, v in sorted(roster.items()) if v[0] not in matched]
        if missing:
            print("Roster raiders not present in the plan:")
            for disp, cls, role in missing:
                print(f"  - {disp} ({cls}, {role})")

    if args.dry_run:
        print("\nDry run — no plan created.")
        return
    if not (marker_hits or text_hits or icon_hits or fmt_hits):
        print("\nNo changes to make — plan left untouched." if (args.check or args.sync_names)
              else "\nNothing matched — no plan created.")
        return

    headers = {"Cookie": args.cookie} if args.cookie else None
    if args.in_place:
        if not (args.key or args.cookie):
            print("\n--in-place needs --key (access key) or --cookie (login session) "
                  "to prove you own the plan.\n")
            sys.exit(key_help(code))
        target, key = code, args.key
        backup = f"backup_{code}_rev{plan.get('revision', 0)}.json"
        with open(backup, "w", encoding="utf-8") as f:
            json.dump(original_doc, f, ensure_ascii=False)
        print(f"\nOriginal document backed up to {backup}")
        print("Saving to the original plan ...")
    else:
        print("\nCloning plan ...")
        clone = http_json(f"https://raidplan.io/api/plans/{code}/clone", payload={}, headers=headers)
        target, key = clone["plan"]["code"], clone["access_key"]

    payload = {
        "access_key": key,
        "steps": doc["steps"],
        "name": args.name or (plan["name"] if args.in_place else f"{plan['name']} v2"),
        "nodes": doc["nodes"],
        "header_notes": doc.get("header_notes_raw") or "",
        "footer_notes": doc.get("footer_notes_raw") or "",
        "step_notes": step_notes or [""] * doc["steps"],
        "single_note": doc.get("single_note", False),
    }
    try:
        http_json(f"https://raidplan.io/api/plans/{target}", payload=payload, headers=headers)
    except urllib.error.HTTPError as e:
        if args.in_place and e.code in (401, 403):
            print(f"\nSave rejected (HTTP {e.code}) — the key/cookie does not grant edit access "
                  f"to plan '{target}'.\n")
            sys.exit(key_help(target))
        raise

    if args.in_place:
        print(f"Done. Plan updated in place:  https://raidplan.io/plan/{target}")
    else:
        print(f"Done. New plan:  https://raidplan.io/plan/{target}")
        print(f"Edit link (keep private, grants edit access):")
        print(f"  https://raidplan.io/plan/{target}/{key}")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
